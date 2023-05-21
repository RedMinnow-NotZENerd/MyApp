from flask import Flask, render_template, request, redirect, flash
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from passlib.hash import sha256_crypt
from flask_wtf import FlaskForm
from wtforms import StringField, PasswordField, SubmitField
from wtforms.validators import DataRequired
import openpyxl

app = Flask(__name__)
app.config['SECRET_KEY'] = 'e5d37f9036b78a23785e70b89ae09129'

# Configure the LoginManager
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Database file path
DATABASE_PATH = 'C:\\Users\\Sahil\\Downloads\\database.xlsx'

def load_users():
    workbook = openpyxl.load_workbook(DATABASE_PATH)
    worksheet = workbook.active
    max_row = worksheet.max_row

    users = []
    for row in range(2, max_row + 1):
        username = worksheet.cell(row=row, column=1).value
        hashed_password = worksheet.cell(row=row, column=2).value
        user_id = username  # You can set user_id to be the same as the username
        users.append(User(user_id, username, hashed_password))

    workbook.close()
    return users



class User(UserMixin):
    def __init__(self, user_id, username, password):
        self.id = user_id
        self.username = username
        self.password = password
        self.sent_messages = []
        self.received_messages = []


    @staticmethod
    def get(user_id):
        # Retrieve the user from the database
        workbook = openpyxl.load_workbook(DATABASE_PATH)
        worksheet = workbook.active
        max_row = worksheet.max_row

        for row in range(2, max_row + 1):
            username = worksheet.cell(row=row, column=1).value
            hashed_password = worksheet.cell(row=row, column=2).value
            if username == user_id:
                workbook.close()
                return User(user_id, username, hashed_password)

        workbook.close()
        return None


    def verify_password(self, password):
        return sha256_crypt.verify(password, self.password)


class Message:
    def __init__(self, sender, receiver, text):
        self.sender = sender
        self.receiver = receiver
        self.text = text


@login_manager.user_loader
def load_user(user_id):
    return User.get(user_id)


# Form for user login
class LoginForm(FlaskForm):
    username = StringField('Username', validators=[DataRequired()])
    password = PasswordField('Password', validators=[DataRequired()])
    submit = SubmitField('Log In')


# Form for user registration
class RegistrationForm(FlaskForm):
    username = StringField('Username', validators=[DataRequired()])
    password = PasswordField('Password', validators=[DataRequired()])
    confirm_password = PasswordField('Confirm Password', validators=[DataRequired()])
    submit = SubmitField('Sign Up')


@app.route('/')
@login_required
def home():
    return render_template('index.html', current_user=current_user)


@app.route('/index')
@login_required
def index():
    users = load_users()
    return render_template('index.html', users=users)


@app.route('/send_message', methods=['POST'])
@login_required
def send_message():
    recipient_username = request.form.get('recipient_username')
    message_text = request.form.get('message_text')

    # Validate recipient username
    recipient = User.get(recipient_username)
    if not recipient:
        flash('Invalid recipient username')
        return redirect('/index')

    # Create a new message
    message = Message(sender=current_user.username, receiver=recipient_username, text=message_text)

    # Add the message to the sender's sent messages
    current_user.sent_messages.append(message)

    # Add the message to the recipient's received messages
    recipient.received_messages.append(message)  # Add this line to add the message to the recipient's received_messages list

    # Save the changes to the database
    workbook = openpyxl.load_workbook(DATABASE_PATH)
    worksheet = workbook.active

    max_row = worksheet.max_row
    for row in range(2, max_row + 1):
        username = worksheet.cell(row=row, column=1).value
        if username == current_user.username:
            sent_messages_column = worksheet.cell(row=row, column=4)
            if sent_messages_column.value:
                sent_messages_column.value += f"\nTo: {recipient_username}\nMessage: {message_text}"
            else:
                sent_messages_column.value = f"To: {recipient_username}\nMessage: {message_text}"
            break

    for row in range(2, max_row + 1):
        username = worksheet.cell(row=row, column=1).value
        if username == recipient_username:
            received_messages_column = worksheet.cell(row=row, column=3)
            if received_messages_column.value:
                received_messages_column.value += f"\nFrom: {current_user.username}\nMessage: {message_text}"
            else:
                received_messages_column.value = f"From: {current_user.username}\nMessage: {message_text}"
            break

    workbook.save(DATABASE_PATH)
    workbook.close()

    flash('Message sent successfully')
    return redirect('/index')









@app.route('/private_chat')
@login_required
def private_chat():
    workbook = openpyxl.load_workbook(DATABASE_PATH)
    worksheet = workbook.active
    max_row = worksheet.max_row

    users = []
    for row in range(2, max_row + 1):
        username = worksheet.cell(row=row, column=1).value
        users.append(username)

    workbook.close()

    return render_template('private_chat.html', users=users)


@app.route('/received_messages')
@login_required
def received_messages():
    messages = []
    for message in current_user.received_messages:
        if message.receiver == current_user.username:
            sender = User.get(message.sender)
            if sender:
                messages.append((sender.username, message.text))
    return render_template('received_messages.html', messages=messages)



@app.route('/login', methods=['GET', 'POST'])
def login():
    form = LoginForm()
    if form.validate_on_submit():
        user_id = form.username.data
        password = form.password.data
        user = User.get(user_id)
        if user and user.verify_password(password):
            login_user(user)
            flash('Logged in successfully!', 'success')
            return redirect('/index')
        else:
            flash('Invalid username or password', 'error')
            return redirect('/login')

    return render_template('login.html', form=form)


@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect('/login')


@app.route('/signup', methods=['GET', 'POST'])
def signup():
    form = RegistrationForm()
    if form.validate_on_submit():
        # Save user credentials to the database
        username = form.username.data
        password = form.password.data
        confirm_password = form.confirm_password.data
        if password != confirm_password:
            flash('Password and Confirm Password do not match.', 'error')
            return redirect('/signup')

        # Hash the password
        hashed_password = sha256_crypt.hash(password)

        # Save the username and hashed password to the database
        workbook = openpyxl.load_workbook(DATABASE_PATH)
        worksheet = workbook.active

        max_row = worksheet.max_row
        new_row = max_row + 1

        worksheet.cell(row=new_row, column=1).value = username
        worksheet.cell(row=new_row, column=2).value = hashed_password

        workbook.save(DATABASE_PATH)
        workbook.close()

        flash('Account created successfully! You can now log in.', 'success')
        return redirect('/login')

    return render_template('signup.html', form=form)


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5003, debug=True)

       








